import os
import pandas as pd
import xlrd
from openpyxl import load_workbook
from extract_pl_pdf import extract_packing_list_data

# -------- Directories -------- #
dir = "/home/pritom/Desktop/C&A Packing List Extractor/media/84771-030-46-130-001"  

# Extract PDF data
df = extract_packing_list_data(dir)
print(df)

# Ensure directory exists
os.makedirs(dir, exist_ok=True)

# -------- Helpers -------- #
def get_sheet_names(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        return xlrd.open_workbook(file_path).sheet_names()
    else:
        return pd.ExcelFile(file_path).sheet_names

# -------- Validation Function -------- #
# Fallback mapping for sheet names
sheet_map = {"DE": "D", "AT": "A", "BE": "B"}

def validate_excel_file(file_path, pdf_orders):
    filename = os.path.basename(file_path)
    
    try:
        if not pdf_orders:
            print(f"❌ No orders in PDF data: {filename}")
            return None
        
        # Always one order
        order_no = list(pdf_orders.keys())[0]
        expected_countries = pdf_orders[order_no]
        
        sheet_names = get_sheet_names(file_path)
        available_sheets = [s.upper() for s in sheet_names]

        missing_sheets = []

        for country in expected_countries:
            # Direct match
            if country in available_sheets:
                continue
            # Check fallback mapping
            mapped = sheet_map.get(country)
            if mapped and mapped.upper() in available_sheets:
                continue
            # Still missing
            missing_sheets.append(country)

        if missing_sheets:
            print(f"❌ Missing sheets for {order_no}: {missing_sheets} in {filename}")
            return None

        print(f"✅ Excel validated for order {order_no} — all sheets present")
        return order_no, file_path

    except Exception as e:
        print(f"❌ Error processing {filename}: {e}")
        return None

# -------- New: Dynamic weight update -------- #
def update_excel_weights(file_path, pdf_df):
    wb = load_workbook(file_path)

    try:
        # Iterate over all rows in df
        for _, row in pdf_df.iterrows():
            country = row['country_iso']
            net_weight = row['net_weight']
            gross_weight = row['gross_weight']

            # Determine sheet name (fallback if needed)
            sheet_name = country
            if sheet_name not in wb.sheetnames and country in sheet_map:
                sheet_name = sheet_map[country]

            if sheet_name not in wb.sheetnames:
                raise ValueError(f"❌ Sheet for country {country} not found!")

            ws = wb[sheet_name]

            # --- Update weights ---
            update_done = False
            for excel_row in ws.iter_rows():
                for cell in excel_row:
                    if isinstance(cell.value, str):
                        if "total net" in cell.value.lower():
                            next_cell = ws.cell(cell.row + 1, cell.column)
                            if next_cell.value and "weight" in str(next_cell.value).lower():
                                ws.cell(cell.row + 2, cell.column, net_weight)
                                update_done = True
                        if "total gross" in cell.value.lower():
                            next_cell = ws.cell(cell.row + 1, cell.column)
                            if next_cell.value and "weight" in str(next_cell.value).lower():
                                ws.cell(cell.row + 2, cell.column, gross_weight)
                                update_done = True

            if not update_done:
                raise ValueError(f"❌ No cells updated in sheet '{sheet_name}' for country {country}!")

        # Save workbook only if all updates succeeded
        wb.save(file_path)
        print(f"✅ Updated all weights in {os.path.basename(file_path)}")

    except Exception as e:
        print(f"❌ Failed to update Excel file '{os.path.basename(file_path)}': {e}")
        print("⚠️ Update canceled, file not saved.")
        
# -------- Main Driver -------- #
def update_packing_lists():
    if df.empty:
        print("❌ No PDF data available - terminating")
        return

    # Build dictionary of countries for the single order
    pdf_orders = {order_no: group['country_iso'].unique().tolist()
                  for order_no, group in df.groupby('order_no')}

    if len(pdf_orders) != 1:
        print(f"❌ Expected exactly 1 order in PDF, found {len(pdf_orders)}")
        return

    # Get Excel files
    excel_files = [f for f in os.listdir(dir) if f.lower().endswith((".xls", ".xlsx", ".xlsm"))]

    if not excel_files:
        print("❌ No Excel files found - terminating")
        return

    # Always use first file (assume only one)
    file_path = os.path.join(dir, excel_files[0])
    
    if validate_excel_file(file_path, pdf_orders):
        update_excel_weights(file_path, df)  # <-- dynamically update all sheets

# -------- Run Script -------- #
if __name__ == "__main__":
    update_packing_lists()