import os
import subprocess
from openpyxl import load_workbook

input_dir = "/home/pritom/Desktop/C&A Packing List Extractor/input"
output_dir = os.path.join(input_dir, "output")
TARGET_SHEET = "summary sheet"
NEW_ORDER_NUMBER = "123456"  # change this

os.makedirs(output_dir, exist_ok=True)

# ---------------- Convert XLS → XLSX ---------------- #
def convert_xls_to_xlsx(xls_file):
    xlsx_file = xls_file + "x"  # file.xls -> file.xlsx

    cmd = [
        "libreoffice", "--headless",
        "--convert-to", "xlsx", xls_file,
        "--outdir", input_dir
    ]
    subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

    return xlsx_file if os.path.exists(xlsx_file) else None


# ---------------- Replace Order Number in XLSX ---------------- #
def update_order_number(xlsx_path, out_path):
    wb = load_workbook(xlsx_path, data_only=True)
    
    if TARGET_SHEET not in wb.sheetnames:
        print(f"⚠️ Summary sheet not found in: {os.path.basename(xlsx_path)}")
        wb.save(out_path)
        return
    
    ws = wb[TARGET_SHEET]
    replaced = False

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "order number" in cell.value.lower():
                for right in row[row.index(cell) + 1:]:
                    if right.value not in [None, ""]:
                        right.value = NEW_ORDER_NUMBER
                        replaced = True
                        break
                break

    wb.save(out_path)

    if replaced:
        print(f"✅ Updated Order Number → {os.path.basename(out_path)}")
    else:
        print(f"⚠️ Order Number not found → {os.path.basename(out_path)}")


# ---------------- Main Driver ---------------- #
def run():
    for file in os.listdir(input_dir):
        if file.lower().endswith(".xls"):
            xls_path = os.path.join(input_dir, file)
            print(f"📂 Converting: {file}")

            xlsx_path = convert_xls_to_xlsx(xls_path)
            if not xlsx_path:
                print(f"❌ Failed to convert {file}")
                continue

            output_xlsx = os.path.join(output_dir, os.path.basename(xlsx_path))
            update_order_number(xlsx_path, output_xlsx)

            os.remove(xlsx_path)  # temporary XLSX removed

    print("✅ Done! Converted and updated XLS → XLSX")


if __name__ == "__main__":
    run()
