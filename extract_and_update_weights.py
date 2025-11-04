import os
import pandas as pd
import xlrd
from openpyxl import load_workbook, Workbook
import shutil

input_dir = "/home/pritom/Desktop/C&A Packing List Extractor/input/Upload"
out_dir = "/home/pritom/Desktop/C&A Packing List Extractor/output"
TARGET_SHEET = "NL"
NEW_VALUE = 12

os.makedirs(out_dir, exist_ok=True)

def get_sheet_names(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        return xlrd.open_workbook(file_path).sheet_names()
    else:
        return pd.ExcelFile(file_path).sheet_names


def update_weights_in_nl(file_path, output_path):
    try:
        sheet_names = get_sheet_names(file_path)
        nl_sheet = next((s for s in sheet_names if s.strip().upper() == TARGET_SHEET), None)

        if not nl_sheet:
            print(f"⚠️ No 'NL' sheet in {os.path.basename(file_path)} — copied without change")
            shutil.copy(file_path, output_path)
            return

        wb = load_workbook(file_path)
        ws = wb[nl_sheet]

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "total net" in cell.value.lower():
                    next_cell = ws.cell(cell.row + 1, cell.column)
                    if next_cell.value and "weight" in str(next_cell.value).lower():
                        ws.cell(cell.row + 2, cell.column, NEW_VALUE)

                if isinstance(cell.value, str) and "total gross" in cell.value.lower():
                    next_cell = ws.cell(cell.row + 1, cell.column)
                    if next_cell.value and "weight" in str(next_cell.value).lower():
                        ws.cell(cell.row + 2, cell.column, NEW_VALUE)

        wb.save(output_path)
        print(f"✅ Updated weights → saved: {os.path.basename(output_path)}")

    except Exception as e:
        print(f"❌ Error processing {file_path}: {e}")


def main():
    files = [f for f in os.listdir(input_dir) if f.lower().endswith((".xls", ".xlsx", ".xlsm"))]

    if not files:
        print("❌ No Excel files found in input directory")
        return

    print(f"📂 Found {len(files)} files\n")

    for file in files:
        in_path = os.path.join(input_dir, file)
        out_path = os.path.join(out_dir, file)

        update_weights_in_nl(in_path, out_path)


if __name__ == "__main__":
    main()
