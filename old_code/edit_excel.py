import os
from openpyxl import load_workbook
import shutil

input_dir = "/home/pritom/Desktop/C&A Packing List Extractor/input/xlsx"
output_dir = os.path.join(input_dir, "output")
TARGET_SHEET = "Summary Sheet"
NEW_ORDER_NUMBER = "123456"   # <-- change this

os.makedirs(output_dir, exist_ok=True)

def replace_order_number(file_path, output_path):
    wb = load_workbook(file_path, data_only=True, keep_vba=True)

    if TARGET_SHEET not in wb.sheetnames:
        print(f"⚠️ Summary sheet not found in: {os.path.basename(file_path)}")
        print(f"   Available sheets: {', '.join(wb.sheetnames)}")
        return

    ws = wb[TARGET_SHEET]
    replaced = False

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "order number" in cell.value.strip().lower():
                for next_cell in row[row.index(cell) + 1:]:
                    if next_cell.value not in [None, ""]:
                        next_cell.value = NEW_ORDER_NUMBER
                        replaced = True
                        break
                break

    # If nothing replaced, copy original file instead
    if not replaced:
        shutil.copy2(file_path, output_path)
        print(f"⚠️ No Order Number cell found, file copied unchanged: {os.path.basename(output_path)}")
        return

    wb.save(output_path)
    print(f"✅ Updated and saved: {os.path.basename(output_path)}")


def run():
    for file in os.listdir(input_dir):
        if file.lower().endswith((".xlsx", ".xlsm")):
            input_file = os.path.join(input_dir, file)
            output_file = os.path.join(output_dir, file)
            replace_order_number(input_file, output_file)


if __name__ == "__main__":
    run()