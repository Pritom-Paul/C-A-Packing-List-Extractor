import os
import pandas as pd
import xlrd
from openpyxl import load_workbook
from extract_pl_pdf_data import extract_packing_list_data

input_dir = "/home/pritom/Desktop/C&A Packing List Extractor/input/Upload"
output_dir = "/home/pritom/Desktop/C&A Packing List Extractor/output"
TARGET_SHEET = "summary sheet"
df = extract_packing_list_data(input_dir)
print(df)

# Create output directory if it doesn't exist
os.makedirs(output_dir, exist_ok=True)

# -------- Helpers -------- #

def find_value_after_cell(rows, label):
    """Return the first non-empty cell immediately after the cell containing `label`."""
    label = label.lower()
    for row in rows:
        for i, cell in enumerate(row):
            if isinstance(cell, str) and label in cell.strip().lower():
                for val in row[i + 1:]:
                    if val is not None and str(val).strip() != "":
                        return str(val).strip()
    return ""

def get_sheet_names(file_path):
    """Get all sheet names from Excel file."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        return xlrd.open_workbook(file_path).sheet_names()
    else:
        return pd.ExcelFile(file_path).sheet_names

def extract_order_from_excel(file_path):
    """Extract order number from Excel file."""
    try:
        sheet_names = get_sheet_names(file_path)
        target_sheet = next((s for s in sheet_names if s.strip().lower() == TARGET_SHEET.lower()), None)
        
        if not target_sheet:
            print(f"    ⚠️ Summary sheet not found in: {os.path.basename(file_path)}")
            return None
        
        df_sheet = pd.read_excel(file_path, sheet_name=target_sheet, header=None)
        rows = df_sheet.fillna("").values.tolist()
        order_no = find_value_after_cell(rows, "order number")
        
        if not order_no:
            print(f"    ⚠️ Order number not found in summary sheet: {os.path.basename(file_path)}")
            return None
            
        return order_no
        
    except Exception as e:
        print(f"    ❌ Error extracting order number from {os.path.basename(file_path)}: {e}")
        return None

def update_excel_weights(file_path, order_no, country_sheets, df):
    """Update net and gross weights in Excel country sheets with values from dataframe."""
    print(f"  📊 Updating weights for order {order_no}...")
    
    try:
        # Load workbook with openpyxl for minimal changes
        wb = load_workbook(file_path, keep_vba=True)
        updated = False
        
        for country in country_sheets:
            # Find the actual sheet name (case-sensitive)
            sheet_name = next((s for s in wb.sheetnames if s.upper() == country), None)
            
            if not sheet_name:
                print(f"    ❌ Sheet {country} not found in workbook")
                return False  # Terminate if any sheet is missing
            
            # Get weight values from dataframe
            mask = (df['order_no'] == order_no) & (df['country_iso'] == country)
            if not mask.any():
                print(f"    ❌ No data found for {country} in order {order_no}")
                return False  # Terminate if data mismatch
            
            net_weight = df.loc[mask, 'net_weight'].iloc[0]
            gross_weight = df.loc[mask, 'gross_weight'].iloc[0]
            
            ws = wb[sheet_name]
            net_updated = False
            gross_updated = False
            
            # Find and update net weight
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "total net" in cell.value.strip().lower():
                        # Check if cell below contains "Weight"
                        next_cell = ws.cell(row=cell.row + 1, column=cell.column)
                        if next_cell.value and "weight" in str(next_cell.value).lower():
                            # Update the value below "Weight"
                            weight_cell = ws.cell(row=cell.row + 2, column=cell.column)
                            old_net_value = weight_cell.value
                            if old_net_value != net_weight:
                                weight_cell.value = net_weight
                                net_updated = True
                                print(f"    ✅ {country}: Net weight changed from {old_net_value} to {net_weight}")
                            else:
                                print(f"    ℹ️  {country}: Net weight unchanged ({net_weight})")
                                net_updated = True
                            break
            
            # Find and update gross weight
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and "total net" in cell.value.strip().lower():
                        next_cell = ws.cell(row=cell.row + 1, column=cell.column)
                        if next_cell.value and "weight" in str(next_cell.value).lower():
                            weight_cell = ws.cell(row=cell.row + 2, column=cell.column)
                            old_gross_value = weight_cell.value
                            if old_gross_value != gross_weight:
                                weight_cell.value = gross_weight
                                gross_updated = True
                                print(f"    ✅ {country}: Gross weight changed from {old_gross_value} to {gross_weight}")
                            else:
                                print(f"    ℹ️  {country}: Gross weight unchanged ({gross_weight})")
                                gross_updated = True
                            break
            
            if not net_updated:
                print(f"    ❌ Failed to find net weight cell for {country}")
                return False
                
            if not gross_updated:
                print(f"    ❌ Failed to find gross weight cell for {country}")
                return False
            
            updated = True
        
        return updated
        
    except Exception as e:
        print(f"  ❌ Error updating Excel file: {e}")
        return False

# -------- Validation Functions -------- #

def validate_excel_file(file_path, pdf_orders):
    """Validate a single Excel file against PDF data."""
    filename = os.path.basename(file_path)
    
    try:
        # Extract order number from Excel
        excel_order_no = extract_order_from_excel(file_path)
        
        if not excel_order_no:
            print(f"❌ Order number extraction failed: {filename}")
            return None
        
        # Check if order exists in PDF data
        if excel_order_no not in pdf_orders:
            print(f"❌ Order {excel_order_no} not in PDF data: {filename}")
            return None
        
        # Validate country sheets
        sheet_names = get_sheet_names(file_path)
        expected_countries = pdf_orders[excel_order_no]
        available_sheets = [s.upper() for s in sheet_names]
        missing_sheets = [c for c in expected_countries if c not in available_sheets]
        
        if missing_sheets:
            print(f"❌ Missing sheets for {excel_order_no}: {missing_sheets} in {filename}")
            return None
        
        print(f"✅ Validated {excel_order_no}: {filename}")
        return excel_order_no, file_path
            
    except Exception as e:
        print(f"❌ Error processing {filename}: {e}")
        return None

# -------- Main Driver -------- #

def update_packing_lists():
    """Main function to validate Excel files against PDF data and update weights."""
    if df.empty:
        print("❌ No PDF data available - terminating")
        return
    
    # Create order-country mapping from PDF data
    pdf_orders = {}
    for order_no, group in df.groupby('order_no'):
        pdf_orders[order_no] = group['country_iso'].unique().tolist()
    
    print(f"🔍 Validating {len(pdf_orders)} orders from PDF data...")
    
    # Process Excel files
    excel_files = [f for f in os.listdir(input_dir) if f.lower().endswith((".xls", ".xlsx", ".xlsm"))]
    
    if not excel_files:
        print("❌ No Excel files found - terminating")
        return
    
    print(f"📁 Found {len(excel_files)} Excel files")
    
    # Track which orders we found
    found_orders = set()
    validated_files = []
    
    # First pass: Quick check files with order number in filename (efficiency)
    for order_no in pdf_orders:
        matching_files = [f for f in excel_files if order_no in f]
        if matching_files:
            found_orders.add(order_no)
            print(f"⚡ Quick match found for order {order_no}")
    
    # Second pass: Check all Excel files for order numbers in summary sheets
    for file in excel_files:
        file_path = os.path.join(input_dir, file)
        result = validate_excel_file(file_path, pdf_orders)
        if result:
            order_no, file_path = result
            found_orders.add(order_no)
            validated_files.append((order_no, file_path))
    
    # Check if we found all PDF orders
    missing_orders = set(pdf_orders.keys()) - found_orders
    if missing_orders:
        print(f"❌ Missing Excel files for orders: {missing_orders}")
        print("❌ Process terminated - not all orders have matching Excel files")
        return
    
    # Update Excel files with weights from dataframe
    updated_count = 0
    for order_no, file_path in validated_files:
        country_sheets = pdf_orders[order_no]
        filename = os.path.basename(file_path)
        output_path = os.path.join(output_dir, filename)
        
        print(f"\n🔄 Processing: {filename}")
        
        # Try to update the file
        updated = update_excel_weights(file_path, order_no, country_sheets, df)
        
        if updated:
            # Save the updated workbook
            wb = load_workbook(file_path, keep_vba=True)
            wb.save(output_path)
            print(f"  💾 Successfully saved: {filename}")
            updated_count += 1
        else:
            print(f"  ❌ Failed to update weights for {filename}")
            print("❌ Process terminated due to update failure")
            return
    
    print(f"\n🎯 Process completed successfully!")
    print(f"✅ Updated {updated_count} Excel files")
    print(f"📁 Output directory: {output_dir}")

if __name__ == "__main__":
    update_packing_lists()